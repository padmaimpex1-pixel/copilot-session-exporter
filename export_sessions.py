"""
Export all Copilot CLI session prompts and assistant responses to an Excel workbook.

Usage:
    python export_sessions.py [output.xlsx]

Output columns:
    Session ID, Session Start (UTC), Workspace, Interaction #,
    Turn #, Timestamp (UTC), Role, Message
"""

import json
import os
import sys
from datetime import datetime, timezone
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.exit("openpyxl is required. Install it with:  pip install openpyxl")

SESSION_STATE_DIR = Path.home() / ".copilot" / "session-state"
DEFAULT_OUTPUT = Path.cwd() / "copilot_sessions.xlsx"

# ---------------------------------------------------------------------------
# Styles
# ---------------------------------------------------------------------------
HEADER_FILL = PatternFill("solid", fgColor="2D3B55")
HEADER_FONT = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
USER_FILL   = PatternFill("solid", fgColor="EAF2FF")
ASST_FILL   = PatternFill("solid", fgColor="F0FFF0")
THIN_BORDER = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)
WRAP = Alignment(wrap_text=True, vertical="top")
TOP  = Alignment(vertical="top")


def utc(ts: str) -> str:
    """Return a human-readable UTC timestamp string."""
    try:
        dt = datetime.fromisoformat(ts.replace("Z", "+00:00"))
        return dt.astimezone(timezone.utc).strftime("%Y-%m-%d %H:%M:%S")
    except Exception:
        return ts


def load_session(session_dir: Path) -> dict | None:
    """Parse events.jsonl from a session folder and return structured data."""
    events_file = session_dir / "events.jsonl"
    if not events_file.exists():
        return None

    meta_file = session_dir / "vscode.metadata.json"
    workspace = ""
    if meta_file.exists():
        try:
            meta = json.loads(meta_file.read_text(encoding="utf-8"))
            wf = meta.get("workspaceFolder", {})
            workspace = wf.get("folderPath", "") if isinstance(wf, dict) else str(wf)
        except Exception:
            pass

    session_start = ""
    interactions: dict[str, list[dict]] = {}   # interactionId -> list of messages
    current_interaction = "__no_interaction__"

    try:
        lines = events_file.read_text(encoding="utf-8", errors="replace").splitlines()
    except Exception:
        return None

    for line in lines:
        line = line.strip()
        if not line:
            continue
        try:
            event = json.loads(line)
        except json.JSONDecodeError:
            continue

        etype = event.get("type", "")
        data  = event.get("data", {})
        ts    = event.get("timestamp", "")

        if etype == "session.start":
            session_start = utc(ts)
            continue

        # Track current interaction grouping
        if etype == "user.message":
            iid = data.get("interactionId", current_interaction)
            current_interaction = iid
            content = data.get("content", "").strip()
            if not content:
                content = data.get("transformedContent", "").strip()
            if iid not in interactions:
                interactions[iid] = []
            interactions[iid].append({
                "role": "user",
                "content": content,
                "timestamp": utc(ts),
            })

        elif etype == "assistant.message":
            iid = data.get("interactionId", current_interaction)
            content = data.get("content", "").strip()
            model   = data.get("model", "")
            turn    = data.get("turnId", "")
            if iid not in interactions:
                interactions[iid] = []
            interactions[iid].append({
                "role": f"assistant ({model})" if model else "assistant",
                "content": content,
                "timestamp": utc(ts),
                "turn": turn,
            })

    if not interactions:
        return None

    return {
        "session_id": session_dir.name,
        "session_start": session_start,
        "workspace": workspace,
        "interactions": interactions,
    }


def write_xlsx(sessions: list[dict], output_path: Path) -> None:
    wb = openpyxl.Workbook()

    # -----------------------------------------------------------------------
    # Sheet 1: Flat message log
    # -----------------------------------------------------------------------
    ws = wb.active
    ws.title = "Messages"

    headers = [
        "Session ID", "Session Start (UTC)", "Workspace",
        "Interaction #", "Turn #", "Timestamp (UTC)", "Role", "Message",
    ]
    ws.append(headers)
    for col_idx, _ in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = TOP
        cell.border = THIN_BORDER

    row_num = 2
    for session in sessions:
        sid       = session["session_id"]
        sstart    = session["session_start"]
        workspace = session["workspace"]

        for interaction_num, (iid, messages) in enumerate(session["interactions"].items(), start=1):
            for msg in messages:
                role    = msg["role"]
                content = msg["content"]
                ts      = msg["timestamp"]
                turn    = msg.get("turn", "")

                fill = USER_FILL if msg["role"] == "user" else ASST_FILL

                values = [sid, sstart, workspace, interaction_num, turn, ts, role, content]
                ws.append(values)
                for col_idx, _ in enumerate(values, start=1):
                    cell = ws.cell(row=row_num, column=col_idx)
                    cell.fill = fill
                    cell.border = THIN_BORDER
                    cell.alignment = WRAP if col_idx == len(values) else TOP
                row_num += 1

    # Column widths
    col_widths = [38, 20, 45, 14, 8, 20, 24, 80]
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    # -----------------------------------------------------------------------
    # Sheet 2: Summary (one row per session)
    # -----------------------------------------------------------------------
    ws2 = wb.create_sheet("Summary")
    sum_headers = [
        "Session ID", "Session Start (UTC)", "Workspace",
        "# Interactions", "# User Messages", "# Assistant Messages",
    ]
    ws2.append(sum_headers)
    for col_idx, _ in enumerate(sum_headers, start=1):
        cell = ws2.cell(row=1, column=col_idx)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = TOP
        cell.border = THIN_BORDER

    for r, session in enumerate(sessions, start=2):
        all_msgs = [m for msgs in session["interactions"].values() for m in msgs]
        user_count = sum(1 for m in all_msgs if m["role"] == "user")
        asst_count = sum(1 for m in all_msgs if m["role"] != "user")
        row = [
            session["session_id"],
            session["session_start"],
            session["workspace"],
            len(session["interactions"]),
            user_count,
            asst_count,
        ]
        ws2.append(row)
        for col_idx, _ in enumerate(row, start=1):
            cell = ws2.cell(row=r, column=col_idx)
            cell.border = THIN_BORDER
            cell.alignment = TOP

    for i, w in enumerate([38, 20, 45, 14, 18, 22], start=1):
        ws2.column_dimensions[get_column_letter(i)].width = w
    ws2.freeze_panes = "A2"
    ws2.auto_filter.ref = ws2.dimensions

    wb.save(output_path)


def main() -> None:
    output_path = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_OUTPUT

    if not SESSION_STATE_DIR.exists():
        sys.exit(f"Session state directory not found: {SESSION_STATE_DIR}")

    print(f"Scanning sessions in: {SESSION_STATE_DIR}")
    session_dirs = sorted(
        [d for d in SESSION_STATE_DIR.iterdir() if d.is_dir()],
        key=lambda d: (d / "events.jsonl").stat().st_mtime if (d / "events.jsonl").exists() else 0,
        reverse=True,
    )

    sessions = []
    for d in session_dirs:
        data = load_session(d)
        if data:
            sessions.append(data)
            total_msgs = sum(len(v) for v in data["interactions"].values())
            print(f"  [OK] {d.name[:8]}... - {len(data['interactions'])} interactions, {total_msgs} messages")
        else:
            print(f"  [--] {d.name[:8]}... - skipped (no messages)")

    if not sessions:
        sys.exit("No sessions with messages found.")

    write_xlsx(sessions, output_path)
    total_interactions = sum(len(s["interactions"]) for s in sessions)
    total_messages     = sum(len(m) for s in sessions for m in s["interactions"].values())
    print(f"\nExported {len(sessions)} sessions / {total_interactions} interactions / {total_messages} messages")
    print(f"Output:  {output_path}")


if __name__ == "__main__":
    main()
