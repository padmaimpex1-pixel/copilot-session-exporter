"""
Query the copilot_sessions.xlsx export produced by export_sessions.py.

Subcommands
-----------
sessions   List all sessions (summary sheet).
show       Print every message in one or more sessions.
search     Full-text search across messages.
stats      Aggregate statistics (message counts, model usage, etc.).

Examples
--------
  python query_sessions.py sessions
  python query_sessions.py sessions --since 2026-07-01
  python query_sessions.py show 34b88269
  python query_sessions.py search "angular"
  python query_sessions.py search "angular" --role user --since 2026-07-01 --limit 20
  python query_sessions.py search "error" --output results.xlsx
  python query_sessions.py stats
  python query_sessions.py stats --since 2026-07-11 --workspace Myapp
"""

import argparse
import re
import sys
from pathlib import Path

try:
    import pandas as pd
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError as e:
    sys.exit(f"Missing dependency: {e}\nInstall with:  pip install pandas openpyxl")

DEFAULT_XLS = Path.cwd() / "copilot_sessions.xlsx"
MAX_CELL    = 32767   # Excel single-cell character limit

# ── terminal colours (no external deps) ────────────────────────────────────────────
BOLD  = "\033[1m"
CYAN  = "\033[96m"
GREEN = "\033[92m"
YELLOW= "\033[93m"
RESET = "\033[0m"

def _c(text, colour):
    return f"{colour}{text}{RESET}" if sys.stdout.isatty() else str(text)


# ── helpers ──────────────────────────────────────────────────────────────────────────────

def load_xlsx(path: Path) -> tuple[pd.DataFrame, pd.DataFrame]:
    """Return (messages_df, summary_df). Raises SystemExit if file missing."""
    if not path.exists():
        sys.exit(f"File not found: {path}\nRun export_sessions.py first.")
    msgs = pd.read_excel(path, sheet_name="Messages", dtype=str).fillna("")
    summ = pd.read_excel(path, sheet_name="Summary",  dtype=str).fillna("")
    # Normalise column names to lowercase-underscore
    msgs.columns = [c.strip().lower().replace(" ", "_").replace("#", "num").replace("(utc)", "").strip("_") for c in msgs.columns]
    summ.columns = [c.strip().lower().replace(" ", "_").replace("#", "num").replace("(utc)", "").strip("_") for c in summ.columns]
    return msgs, summ


def apply_global_filters(df: pd.DataFrame, args: argparse.Namespace) -> pd.DataFrame:
    ts_col = "timestamp" if "timestamp" in df.columns else None
    if ts_col and getattr(args, "since", None):
        df = df[df[ts_col] >= args.since]
    if ts_col and getattr(args, "until", None):
        df = df[df[ts_col] <= args.until + " 23:59:59"]
    if getattr(args, "workspace", None) and "workspace" in df.columns:
        df = df[df["workspace"].str.contains(args.workspace, case=False, na=False)]
    return df


def truncate(text: str, width: int = 120) -> str:
    text = text.replace("\n", " ")
    return text[:width] + "\u2026" if len(text) > width else text


def print_table(df: pd.DataFrame, cols: list[str], widths: list[int]) -> None:
    fmt = "  ".join(f"{{:<{w}}}" for w in widths)
    header = fmt.format(*[c.upper() for c in cols])
    print(_c(header, BOLD))
    print("-" * (sum(widths) + 2 * (len(widths) - 1)))
    for _, row in df.iterrows():
        values = [truncate(str(row.get(c, "")), w) for c, w in zip(cols, widths)]
        print(fmt.format(*values))


def save_xlsx(df: pd.DataFrame, path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Results"
    header_fill = PatternFill("solid", fgColor="2D3B55")
    header_font = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
    border = Border(
        left=Side(style="thin", color="CCCCCC"),   right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),    bottom=Side(style="thin", color="CCCCCC"),
    )
    user_fill = PatternFill("solid", fgColor="EAF2FF")
    asst_fill = PatternFill("solid", fgColor="F0FFF0")
    wrap  = Alignment(wrap_text=True, vertical="top")
    top   = Alignment(vertical="top")

    ws.append(list(df.columns))
    for ci, _ in enumerate(df.columns, 1):
        c = ws.cell(1, ci)
        c.font, c.fill, c.alignment, c.border = header_font, header_fill, top, border

    for ri, (_, row) in enumerate(df.iterrows(), 2):
        role = str(row.get("role", ""))
        fill = user_fill if role == "user" else asst_fill
        for ci, val in enumerate(row, 1):
            text = str(val)[:MAX_CELL]
            c = ws.cell(ri, ci, value=text)
            c.fill, c.border = fill, border
            c.alignment = wrap if ci == len(df.columns) else top

    col_widths = {"session_id": 36, "session_start": 20, "workspace": 40,
                  "interaction_num": 14, "turn_num": 8, "timestamp": 20,
                  "role": 22, "message": 80}
    for ci, col in enumerate(df.columns, 1):
        ws.column_dimensions[get_column_letter(ci)].width = col_widths.get(col, 18)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    wb.save(path)
    print(f"Saved {len(df)} rows to {path}")


# ── subcommands ─────────────────────────────────────────────────────────────────────────

def cmd_sessions(args):
    _, summ = load_xlsx(args.file)
    if getattr(args, "since", None) and "session_start" in summ.columns:
        summ = summ[summ["session_start"] >= args.since]
    if getattr(args, "until", None) and "session_start" in summ.columns:
        summ = summ[summ["session_start"] <= args.until + " 23:59:59"]
    if getattr(args, "workspace", None) and "workspace" in summ.columns:
        summ = summ[summ["workspace"].str.contains(args.workspace, case=False, na=False)]
    if getattr(args, "limit", None):
        summ = summ.head(args.limit)
    if getattr(args, "output", None):
        save_xlsx(summ, Path(args.output)); return

    print(_c(f"\n{len(summ)} session(s)\n", CYAN))
    cols   = ["session_id", "session_start", "workspace", "num_interactions", "num_user_messages", "num_assistant_messages"]
    widths = [10, 20, 40, 16, 18, 22]
    summ = summ.copy()
    summ["session_id"] = summ["session_id"].str[:8] + "\u2026"
    print_table(summ, cols, widths)


def cmd_show(args):
    msgs, _ = load_xlsx(args.file)
    df = msgs[msgs["session_id"].str.startswith(args.session_id)]
    if df.empty:
        sys.exit(f"No session found with prefix: {args.session_id}")
    df = apply_global_filters(df, args)
    if getattr(args, "limit", None):
        df = df.head(args.limit)
    if getattr(args, "output", None):
        save_xlsx(df, Path(args.output)); return

    print(_c(f"\n{len(df)} message(s) in session {args.session_id}\u2026\n", CYAN))
    for _, row in df.iterrows():
        role = row.get("role", "")
        col  = GREEN if role == "user" else YELLOW
        ts   = row.get("timestamp", "")
        print(_c(f"[{ts}] {role}", col))
        print(row.get("message", ""))
        print()


def cmd_search(args):
    msgs, _ = load_xlsx(args.file)
    df = apply_global_filters(msgs, args)

    if getattr(args, "role", None):
        if args.role == "user":
            df = df[df["role"] == "user"]
        else:
            df = df[~(df["role"] == "user")]

    pattern = re.compile(args.keyword, re.IGNORECASE)
    df = df[df["message"].str.contains(pattern, na=False)]

    if getattr(args, "limit", None):
        df = df.head(args.limit)

    if getattr(args, "output", None):
        save_xlsx(df, Path(args.output)); return

    print(_c(f"\n{len(df)} match(es) for '{args.keyword}'\n", CYAN))
    for _, row in df.iterrows():
        sid  = str(row.get("session_id", ""))[:8]
        ts   = row.get("timestamp", "")
        role = row.get("role", "")
        msg  = row.get("message", "")
        col  = GREEN if role == "user" else YELLOW
        snippet = truncate(msg, 200)
        snippet = re.sub(f"({re.escape(args.keyword)})", _c(r"\1", BOLD), snippet, flags=re.IGNORECASE)
        print(f"{_c(sid, CYAN)}  {ts}  {_c(role, col)}")
        print(f"  {snippet}")
        print()


def cmd_stats(args):
    msgs, summ = load_xlsx(args.file)
    msgs = apply_global_filters(msgs, args)

    total_sessions     = summ.shape[0]
    total_interactions = pd.to_numeric(summ.get("num_interactions", pd.Series(dtype=str)), errors="coerce").sum()
    total_messages     = len(msgs)
    user_msgs          = (msgs["role"] == "user").sum()
    asst_msgs          = (msgs["role"] != "user").sum()

    model_counts = (
        msgs[msgs["role"] != "user"]["role"]
        .str.extract(r"\((.+)\)")[0]
        .value_counts()
    )

    msgs = msgs.copy()
    msgs["_date"] = msgs["timestamp"].str[:10]
    msgs_per_day = msgs.groupby("_date").size().sort_index(ascending=False).head(10)

    top_workspaces = msgs[msgs["workspace"] != ""].groupby("workspace").size().sort_values(ascending=False).head(5)

    print(_c("\n=== Stats ===\n", BOLD + CYAN))
    print(f"  Sessions         : {total_sessions}")
    print(f"  Interactions     : {int(total_interactions)}")
    print(f"  Total messages   : {total_messages}")
    print(f"  User messages    : {user_msgs}")
    print(f"  Assistant msgs   : {asst_msgs}")

    if not model_counts.empty:
        print(_c("\n  Model usage:", BOLD))
        for model, count in model_counts.items():
            print(f"    {model:<35} {count}")

    if not msgs_per_day.empty:
        print(_c("\n  Messages per day (last 10):", BOLD))
        for date, count in msgs_per_day.items():
            bar = "#" * min(int(count / max(msgs_per_day) * 30), 30)
            print(f"    {date}  {bar:<30} {count}")

    if not top_workspaces.empty:
        print(_c("\n  Top workspaces:", BOLD))
        for ws, count in top_workspaces.items():
            print(f"    {truncate(ws, 50):<52} {count} msgs")
    print()


# ── entry point ──────────────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        prog="query_sessions",
        description="Query copilot_sessions.xlsx",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__,
    )
    parser.add_argument("--file", default=str(DEFAULT_XLS), type=Path, metavar="FILE",
                        help="Path to xlsx (default: copilot_sessions.xlsx)")

    subs = parser.add_subparsers(dest="cmd", required=True)

    def add_filters(p, include_output=True, include_limit=True):
        p.add_argument("--since",     metavar="YYYY-MM-DD", help="Include rows on or after this date")
        p.add_argument("--until",     metavar="YYYY-MM-DD", help="Include rows on or before this date")
        p.add_argument("--workspace", metavar="TEXT",        help="Filter by workspace substring (case-insensitive)")
        if include_limit:
            p.add_argument("--limit",  type=int, metavar="N",    help="Max rows to show")
        if include_output:
            p.add_argument("--output", metavar="FILE",           help="Save results to a new xlsx")

    p_sessions = subs.add_parser("sessions", help="List all sessions")
    add_filters(p_sessions)

    p_show = subs.add_parser("show", help="Print all messages for a session")
    p_show.add_argument("session_id", help="Session ID prefix (first 8+ chars)")
    add_filters(p_show)

    p_search = subs.add_parser("search", help="Full-text search across messages")
    p_search.add_argument("keyword", help="Keyword or regex pattern")
    p_search.add_argument("--role", choices=["user", "assistant"],
                          help="Restrict to user or assistant messages")
    add_filters(p_search)

    p_stats = subs.add_parser("stats", help="Aggregate statistics")
    add_filters(p_stats, include_output=False, include_limit=False)

    args = parser.parse_args()
    dispatch = {"sessions": cmd_sessions, "show": cmd_show,
                "search": cmd_search,    "stats": cmd_stats}
    dispatch[args.cmd](args)


if __name__ == "__main__":
    main()
